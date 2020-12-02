#! /usr/bin/env python
# -*- coding: UTF-8 -*-
import sys
import os
import time
import random
from PyQt5.Qt import *
from lottery_win import *
import openpyxl
import wordcloud

class lotteryMain(QMainWindow, Ui_MainWindow):

    def __init__(self, parent=None):
        super(lotteryMain, self).__init__(parent)
        self.setupUi(self)
        self._register_callbacks()
        self.peopleDat = None
        self.luckyPeople = 0
        self._show_default_board()

    def _register_callbacks(self):
        self.pushButton_load.clicked.connect(self.callbackDoLoad)
        self.pushButton_go.clicked.connect(self.callbackDoGo)

    def _show_default_board(self):
        picFile = os.path.join(os.getcwd(), 'people_cloud_blank.png')
        picObj = QtGui.QPixmap(picFile).scaled(self.label_showPeople.width(), self.label_showPeople.height())
        self.label_showPeople.setPixmap(picObj)

    def callbackDoLoad(self):
        # Select xlsx file
        self.srcXlsx, dummyType = QtWidgets.QFileDialog.getOpenFileName(self, u"Browse File", os.getcwd(), "All Files(*);;Source Files(*.xlsx)")
        self.lineEdit_srcXlsx.setText(self.srcXlsx)

        # Get people list from xlsx
        wb = openpyxl.load_workbook(self.srcXlsx)
        ws = wb.active

        # Save all people names in txt
        txt = u""
        for i in range(1, ws.max_row):
            txt += " " + ws.cell(row=i, column=3).value

        # Generate word cloud picture by txt
        wc = wordcloud.WordCloud(width = 700,
                                 height = 297,
                                 font_path = "msyh.ttc",
                                 max_words = ws.max_row,
                                 background_color = "white")
        wc.generate(txt)
        picFile = os.path.join(os.getcwd(), 'people_cloud.png')
        wc.to_file(picFile)

        # Show word cloud picture
        picObj = QtGui.QPixmap(picFile).scaled(self.label_showPeople.width(), self.label_showPeople.height())
        self.label_showPeople.setPixmap(picObj)

        self.peopleDat = wb

    def callbackDoGo(self):
        self.textEdit_luckyPeopleBoard.clear()

        if self.peopleDat == None:
            QMessageBox.about(self, u"Info", u"Please select people data file (.xlsx)! " )
            return

        wb = self.peopleDat
        ws = wb.active

        if self.lineEdit_luckyPeople.text() != "":
            self.luckyPeople = int(self.lineEdit_luckyPeople.text())

        # Check lucky people number
        if self.luckyPeople == 0 or self.luckyPeople > ws.max_row:
            QMessageBox.about(self, u"Info", u"Please set proper lucky people! " )
            return

        # Get lucky people
        luckyPeopleIdx = []
        while len(luckyPeopleIdx) < self.luckyPeople:
            idx = random.randint(1, ws.max_row)
            if idx not in luckyPeopleIdx:
                luckyPeopleIdx.append(idx)
                #print(idx, ws.cell(row=idx, column=3).value)
                strIdx = str(idx)
                while len(strIdx) < len(str(ws.max_row)):
                    strIdx = '0' + strIdx 
                self.textEdit_luckyPeopleBoard.append(strIdx + " -> " + ws.cell(row=idx, column=3).value + "\r\n")
                time.sleep(1)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainWin = lotteryMain()
    mainWin.setWindowTitle(u"pzh-lottery v1.0")
    mainWin.show()

    sys.exit(app.exec_())

